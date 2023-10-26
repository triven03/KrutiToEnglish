# importing openpyxl module
import openpyxl
import xlwt
from xlwt import Workbook

# Give the location of the file
inputFile = "C:\\File\\inputFile.xlsx"
outputFile = "C:\\File\\outputFile.xls"

# workbook object is created
inpt_obj = openpyxl.load_workbook(inputFile)

inptsheet_obj = inpt_obj.active

# Workbook is created
wb = Workbook()

sheet1 = wb.add_sheet('Sheet 1')


def get_maximum_rows(sheet_obj):
    rows = 0
    for max_row, row in enumerate(sheet_obj, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


m_row = get_maximum_rows(inptsheet_obj)
# print(m_row)
# Loop will print all values
# of first column
sdata = []
for i in range(2, m_row + 1):
    surname = inptsheet_obj.cell(row=i, column=1).value
    name = inptsheet_obj.cell(row=i, column=2).value
    father = inptsheet_obj.cell(row=i, column=3).value
    village = inptsheet_obj.cell(row=i, column=4).value
    data = {
        "Surname": surname,
        "Member": name,
        "Father": father,
        "Village": village
    }
    sdata.append(data)

# print(sdata)

style = xlwt.easyxf('font: bold 1')

# Specifying column
sheet1.write(0, 0, 'Surname', style)
sheet1.write(0, 1, 'Name', style)
sheet1.write(0, 2, 'Father', style)
sheet1.write(0, 3, 'Village', style)


i = 1
sizeofList = len(sdata)
while i < sizeofList:
    # print(sdata[i]["Surname"])
    sheet1.write(i, 0, sdata[i]["Surname"])
    sheet1.write(i, 1, sdata[i]["Member"])
    sheet1.write(i, 2, sdata[i]["Father"])
    sheet1.write(i, 3, sdata[i]["Village"])
    i += 1
# sheet1.write(1, 0, 'ISBT DEHRADUN')
# sheet1.write(2, 0, 'SHASTRADHARA')
# sheet1.write(3, 0, 'CLEMEN TOWN')
# sheet1.write(4, 0, 'RAJPUR ROAD')
# sheet1.write(5, 0, 'CLOCK TOWER')
# sheet1.write(0, 1, 'ISBT DEHRADUN')
# sheet1.write(0, 2, 'SHASTRADHARA')
# sheet1.write(0, 3, 'CLEMEN TOWN')
# sheet1.write(0, 4, 'RAJPUR ROAD')
# sheet1.write(0, 5, 'CLOCK TOWER')

wb.save(outputFile)
