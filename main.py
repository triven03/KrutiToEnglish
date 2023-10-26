import math
from tkinter import *
from PIL import ImageTk, Image
from tkinter import messagebox
from googletrans import Translator
import openpyxl
import xlwt
from xlwt import Workbook, Formula
from operator import itemgetter

# Give the location of the file
inputFile = "C:\\File\\inputFile.xlsx"
namesFile = "C:\\File\\names.xlsx"
outputFile = "C:\\File\\outputFile.xls"
allNamesFile = "C:\\File\\AllNames.xlsx"

# workbook object is created
inpt_obj = openpyxl.load_workbook(inputFile)
inptsheet_obj = inpt_obj.active

names_obj = openpyxl.load_workbook(namesFile)
namessheet_obj = names_obj.active

AllNames_obj = openpyxl.load_workbook(allNamesFile)
AllNamesSheet_obj = AllNames_obj.active

# Workbook is created
wb = Workbook()
sheet1 = wb.add_sheet('Sheet 1')
# sheet data
sheetdata = []
sampleName = []
englishSurname = []
englishName = []
englishFather = []
englishVillage = []
ledger = []
AllMale = []
AllFemale = []


#


class KrutidevToUnicode:
    CHARS_KD = [
        "ñ", "Q+Z", "sas", "aa", ")Z", "ZZ", "‘", "’", "“", "”",

        "å", "ƒ", "„", "…", "†", "‡", "ˆ", "‰", "Š", "‹",

        "¶+", "d+", "[+k", "[+", "x+", "T+", "t+", "M+", "<+", "Q+", ";+", "j+", "u+",
        "Ùk", "Ù", "ä", "–", "—", "é", "™", "=kk", "f=k",

        "à", "á", "â", "ã", "ºz", "º", "í", "{k", "{", "=", "«",
        "Nî", "Vî", "Bî", "Mî", "<î", "|", "K", "}",
        "J", "Vª", "Mª", "<ªª", "Nª", "Ø", "Ý", "nzZ", "æ", "ç", "Á", "xz", "#", ":",

        "v‚", "vks", "vkS", "vk", "v", "b±", "Ã", "bZ", "b", "m", "Å", ",s", ",", "_",

        "ô", "d", "Dk", "D", "[k", "[", "x", "Xk", "X", "Ä", "?k", "?", "³",
        "pkS", "p", "Pk", "P", "N", "t", "Tk", "T", ">", "÷", "¥",

        "ê", "ë", "V", "B", "ì", "ï", "M+", "<+", "M", "<", ".k", ".",
        "r", "Rk", "R", "Fk", "F", ")", "n", "/k", "èk", "/", "Ë", "è", "u", "Uk", "U",

        "i", "Ik", "I", "Q", "¶", "c", "Ck", "C", "Hk", "H", "e", "Ek", "E",
        ";", "¸", "j", "y", "Yk", "Y", "G", "o", "Ok", "O",
        "'k", "'", "\"k", "\"", "l", "Lk", "L", "g",

        "È", "z",
        "Ì", "Í", "Î", "Ï", "Ñ", "Ò", "Ó", "Ô", "Ö", "Ø", "Ù", "Ük", "Ü",

        "‚", "ks", "kS", "k", "h", "q", "w", "`", "s", "S",
        "a", "¡", "%", "W", "•", "·", "∙", "·", "~j", "~", "\\", "+", " ः",
        "^", "*", "Þ", "ß", "(", "¼", "½", "¿", "À", "¾", "A", "-", "&", "&", "Œ", "]", "~ ", "@"
    ]

    CHARS_UNICODE = [
        "॰", "QZ+", "sa", "a", "र्द्ध", "Z", "\"", "\"", "'", "'",

        "०", "१", "२", "३", "४", "५", "६", "७", "८", "९",

        "फ़्", "क़", "ख़", "ख़्", "ग़", "ज़्", "ज़", "ड़", "ढ़", "फ़", "य़", "ऱ", "ऩ",
        "त्त", "त्त्", "क्त", "दृ", "कृ", "न्न", "न्न्", "=k", "f=",

        "ह्न", "ह्य", "हृ", "ह्म", "ह्र", "ह्", "द्द", "क्ष", "क्ष्", "त्र", "त्र्",
        "छ्य", "ट्य", "ठ्य", "ड्य", "ढ्य", "द्य", "ज्ञ", "द्व",
        "श्र", "ट्र", "ड्र", "ढ्र", "छ्र", "क्र", "फ्र", "र्द्र", "द्र", "प्र", "प्र", "ग्र", "रु", "रू",

        "ऑ", "ओ", "औ", "आ", "अ", "ईं", "ई", "ई", "इ", "उ", "ऊ", "ऐ", "ए", "ऋ",

        "क्क", "क", "क", "क्", "ख", "ख्", "ग", "ग", "ग्", "घ", "घ", "घ्", "ङ",
        "चै", "च", "च", "च्", "छ", "ज", "ज", "ज्", "झ", "झ्", "ञ",

        "ट्ट", "ट्ठ", "ट", "ठ", "ड्ड", "ड्ढ", "ड़", "ढ़", "ड", "ढ", "ण", "ण्",
        "त", "त", "त्", "थ", "थ्", "द्ध", "द", "ध", "ध", "ध्", "ध्", "ध्", "न", "न", "न्",

        "प", "प", "प्", "फ", "फ्", "ब", "ब", "ब्", "भ", "भ्", "म", "म", "म्",
        "य", "य्", "र", "ल", "ल", "ल्", "ळ", "व", "व", "व्",
        "श", "श्", "ष", "ष्", "स", "स", "स्", "ह",

        "ीं", "्र",
        "द्द", "ट्ट", "ट्ठ", "ड्ड", "कृ", "भ", "्य", "ड्ढ", "झ्", "क्र", "त्त्", "श", "श्",

        "ॉ", "ो", "ौ", "ा", "ी", "ु", "ू", "ृ", "े", "ै",
        "ं", "ँ", "ः", "ॅ", "ऽ", "ऽ", "ऽ", "ऽ", "्र", "्", "?", "़", ":",
        "‘", "’", "“", "”", ";", "(", ")", "{", "}", "=", "।", ".", "-", "µ", "॰", ",", "् ", "/"
    ]

    @staticmethod
    def do_convert(krutidevPart):
        processPart = str(krutidevPart)
        if processPart != "":
            for input_symbol_idx in range(0, len(KrutidevToUnicode.CHARS_KD)):
                idx = 0
                while idx > -1:
                    processPart = processPart.replace(str(KrutidevToUnicode.CHARS_KD[input_symbol_idx]),
                                                      str(KrutidevToUnicode.CHARS_UNICODE[input_symbol_idx]))
                    idx = processPart.find(str(KrutidevToUnicode.CHARS_KD[input_symbol_idx]))

            # Code for Replacing five Special glyphs

            # Code for Glyph1 : ± (reph+anusvAr)

            processPart = processPart.replace(u'±', u"Zं")

            # Glyp2: Æ
            # code for replacing "f" with "ि" and correcting its position too. (moving it one position forward)

            processPart = processPart.replace(u'Æ', u"र्f")

            position_of_i = processPart.find(u'f')
            while position_of_i > -1:
                charecter_next_to_i = processPart[position_of_i + 1]
                charecter_to_be_replaced = u"f" + charecter_next_to_i
                processPart = processPart.replace(charecter_to_be_replaced, charecter_next_to_i + u"ि")
                position_of_i = processPart.find(u'f', position_of_i + 1)

            # Glyph3 & Glyph4: Ç  É
            # code for replacing "fa" with "िं"  and correcting its position too.(moving it two positions forward)

            processPart = processPart.replace(u'Ç', u"fa")
            processPart = processPart.replace(u'É', u"र्fa")

            position_of_i = processPart.find(u'fa')
            while position_of_i > -1:
                charecter_next_to_ip2 = processPart[position_of_i + 2]
                charecter_to_be_replaced = u"fa" + charecter_next_to_ip2
                processPart = processPart.replace(charecter_to_be_replaced, charecter_next_to_ip2 + u"िं")
                position_of_i = processPart.find(u'fa', position_of_i + 1)

            # Glyph5: Ê
            # code for replacing "h" with "ी"  and correcting its position too.(moving it one positions forward)

            processPart = processPart.replace(u'Ê', u"ीZ")

            # End of Code for Replacing four Special glyphs

            # following loop to eliminate 'chhotee ee kee maatraa' on half-letters as a result of above transformation.
            position_of_wrong_ee = processPart.find(u"ि्")
            while position_of_wrong_ee > -1:
                consonent_next_to_wrong_ee = processPart[position_of_wrong_ee + 2]
                charecter_to_be_replaced = u"ि्" + consonent_next_to_wrong_ee
                processPart = processPart.replace(charecter_to_be_replaced, u"्" + consonent_next_to_wrong_ee + u"ि")
                position_of_wrong_ee = processPart.find(u"ि्", position_of_wrong_ee + 2)

            # Eliminating reph "Z" and putting 'half - r' at proper position for this.
            set_of_matras = u"अ आ इ ई उ ऊ ए ऐ ओ औ ा ि ी ु ू ृ े ै ो ौ ं : ँ ॅ"
            position_of_R = processPart.find(u"Z")
            while position_of_R > -1:
                probable_position_of_half_r = position_of_R - 1
                charecter_at_probable_position_of_half_r = processPart[probable_position_of_half_r]
                # trying to find non-maatra position left to current O (ie, half -r).
                while set_of_matras.find(charecter_at_probable_position_of_half_r) >= 0:
                    probable_position_of_half_r = probable_position_of_half_r - 1
                    charecter_at_probable_position_of_half_r = processPart[probable_position_of_half_r]

                charecter_to_be_replaced = processPart[
                                           probable_position_of_half_r: position_of_R]
                new_replacement_string = u"र्" + charecter_to_be_replaced
                charecter_to_be_replaced = charecter_to_be_replaced + u"Z"
                processPart = processPart.replace(charecter_to_be_replaced, new_replacement_string)
                position_of_R = processPart.find("uZ")

        return processPart

    @staticmethod
    def convert_to_unicode(krutidevString):
        unicodeString = ''

        text_size = len(krutidevString)
        sthiti1 = 0
        sthiti2 = 0
        chale_chalo = 1
        max_text_size = 6000

        while chale_chalo == 1:
            sthiti1 = sthiti2

            if sthiti2 < (text_size - max_text_size):
                sthiti2 += max_text_size
                while krutidevString[sthiti2] != ' ':
                    sthiti2 -= 1
            else:
                sthiti2 = text_size
                chale_chalo = 0

            modifiedSubstring = krutidevString[sthiti1:sthiti2]
            unicodeString += KrutidevToUnicode.do_convert(modifiedSubstring)

        return unicodeString.strip()

    @staticmethod
    def toUnicode(word):
        if not len(word):
            return " "

        word = word.replace("-", "")
        word = word.replace("_`", "_")
        return KrutidevToUnicode.convert_to_unicode(word)


k = KrutidevToUnicode()


def get_maximum_rows(obj):
    rows = 0
    for max_row, row in enumerate(obj, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def checkFather(a):
    if a != None:
        return a
    return "फादर"


def addheadres():
    style = xlwt.easyxf('font: bold 1,height 220 ,color dark_red;alignment: wrap True;align: vert centre, horiz centre')
    style1 = xlwt.easyxf('font:height 220 ;alignment: wrap True;align: vert centre, horiz centre')
    # Specifying column
    sheet1.write(0, 0, 'Customer Type', style)
    sheet1.write(0, 1, 'Member Type', style)
    sheet1.write(0, 2, 'Admission No.', style)
    sheet1.write(0, 3, 'Surname', style)
    sheet1.write(0, 4, 'Member Name', style)
    sheet1.write(0, 5, 'Gender', style)
    sheet1.write(0, 6, 'Share Balance', style)
    sheet1.write(0, 7, 'Village', style)
    sheet1.write(0, 8, 'Ledger Folio No.', style)
    sheet1.write(0, 9, 'Admission Date', style)
    sheet1.write(0, 10, 'Date of Birth', style)
    sheet1.write(0, 11, 'Father name', style)
    sheet1.write(0, 12, 'Community', style)
    sheet1.write(0, 13, 'Farmer Type', style)
    sheet1.write(0, 14, 'Mobile No.', style1)
    sheet1.write(0, 15, 'Aadhaar Card No.', style1)
    sheet1.write(0, 16, 'Account Details', style1)

    for x in range(17):
        sheet1.col(x).width = 20 * 256;

    print("Done add Header")


def checksurname(a, b):
    if a != None:
        return a
    return b


def checkAdhar(a):
    if a is not None:
        return a
    return ""


def checkHissa(a):
    if a is not None:
        return a
    return 0


def readsheetdata(inptsheet_obj):
    m_row = get_maximum_rows(inptsheet_obj)
    sheetdata = []
    for i in range(2, m_row + 1):
        name = k.toUnicode(checksurname(inptsheet_obj.cell(row=i, column=1).value, inptsheet_obj.cell(row=i, column=2).value))
        surname = k.toUnicode(
            checksurname(inptsheet_obj.cell(row=i, column=2).value, inptsheet_obj.cell(row=i, column=1).value))
        father = k.toUnicode(checkFather(inptsheet_obj.cell(row=i, column=3).value))
        village = k.toUnicode(checkAdhar(inptsheet_obj.cell(row=i, column=4).value))
        hissa = checkHissa(inptsheet_obj.cell(row=i, column=5).value)
        adhar = checkAdhar(inptsheet_obj.cell(row=i, column=6).value)

        data = {
            "Customer_Type": "",
            "Member_Type": "",
            "Surname": surname,
            "Member": name,
            "Father": father,
            "Village": village,
            "Hissa": hissa,
            "Adhar": adhar,
            "Gender": "",
            "Farmer": "",
            "Ledger": ""
        }
        sheetdata.append(data)
    # print(sheetdata)
    print("Done Read Data & Conver it to Unicode")
    return sheetdata


def savedatatofile(data):
    # i = 1
    addheadres()
    style1 = xlwt.easyxf('font:height 200 ;alignment: wrap True;align: vert centre, horiz centre')
    sizeofList = len(data)
    for i in range(sizeofList):
        # print(sheetdata[i]["Surname"])
        sheet1.write(i + 1, 0, data[i]["Customer_Type"], style1)
        sheet1.write(i + 1, 1, data[i]["Member_Type"], style1)
        sheet1.write(i + 1, 3, data[i]["Surname"], style1)
        sheet1.write(i + 1, 4, data[i]["Member"], style1)
        sheet1.write(i + 1, 5, data[i]["Gender"], style1)
        sheet1.write(i + 1, 6, data[i]["Hissa"], style1)
        sheet1.write(i + 1, 7, data[i]["Village"], style1)
        sheet1.write(i + 1, 8, data[i]["Ledger"], style1)
        sheet1.write(i + 1, 11, data[i]["Father"], style1)
        sheet1.write(i + 1, 13, data[i]["Farmer"], style1)
        sheet1.write(i + 1, 15, data[i]["Adhar"], style1)

    wb.save(outputFile)
    print("Done Save Data")


def addLedgerno():
    print("Done add ledger")


def toEnglish(limit, data):
    translator = Translator()
    tLength = len(data)
    length = tLength

    if tLength >= limit:
        count = math.ceil(tLength / limit)
    else:
        count = 1

    start = 0
    sample = ""
    totalData = ""
    for i in range(count):
        end = min(limit * (i + 1), length)
        # print(end)
        for na in range(start, end):
            if na != tLength - 1:
                sample = sample + str(data[na]) + ","
            else:
                sample = sample + str(data[na])

        # print(sample)
        transliterated_data = translator.translate(str(sample), dest='hi').pronunciation
        totalData = totalData + transliterated_data

        sample = ""
        start = end
    return totalData


def unicodetoenglish(data):
    # my_input = ("I am very happy to be here with you today to receive the Nobel Prize for Peace.")
    # print(my_input)
    print("Converting data to english.....")
    name = []
    surname = []
    father = []
    village = []

    for na in data:
        name.append(na["Member"])
        surname.append(na["Surname"])
        father.append(na["Father"])
        village.append(na["Village"])

    surtotalData = toEnglish(180, surname)
    memtotalData = toEnglish(180, name)
    fathertotalData = toEnglish(150, father)
    # villtotalData = toEnglish(170, village)

    Name = list(map(str.strip, memtotalData.split(',')))
    Surname = list(map(str.strip, surtotalData.split(',')))
    Father = list(map(str.strip, fathertotalData.split(',')))
    # Village = list(map(str.strip, villtotalData.split(',')))

    allEnglish = {
        "Name": Name,
        "Surname": Surname,
        "Father": Father,
        "Village": village
    }

    print("Done unicode to english conversion")
    return allEnglish


def getLast(name, digit):
    if (len(name) > digit):
        return name[-digit:]
    return name


def getGender(name):
    upperWord = name.upper();
    last2 = getLast(upperWord, 2)
    last3 = getLast(upperWord, 3)
    last4 = getLast(upperWord, 4)
    last5 = getLast(upperWord, 5)

    if last2 == 'HA' or last2 == 'RU' or last2 == 'VA' or last2 == 'LU' or last3 == 'RAM' or last3 == 'DAS' or last3 == 'LAL' or last3 == 'WAR' or last3 == 'SHU' or last3 == 'KHU' or last3 == 'DRA' or last5 == "CHAND" or last4 == "KHAN":
        return "Male"

    elif last2 == 'IN' or last2 == 'YA' or last3 == 'BAI' or last3 == 'ATI' or last3 == 'ILA' or last3 == 'IYA' or last3 == "ARI" or last4 == "DEVI" or "BAI" in upperWord or "KUMARI" in upperWord or "MINA" == upperWord:
        return "Female"

    elif "KUMAR" in upperWord or "RAM" in upperWord or "DAS" in upperWord or "SINGH" in upperWord or "SING" in upperWord or "PRASAD" in upperWord or "LAL" in upperWord:
        return "Male"

    elif any(name in s for s in AllMale):
        return "Male"

    elif any(name in s for s in AllFemale):
        return "Female"

    return "Male"


def getFarmer(ammount):
    ammount = int(ammount)
    if 0 <= ammount <= 999:
        return "Small Or Marginal"
    elif 1000 <= ammount <= 4999:
        return "Medium"
    return "Big"


def getCustomer_Type(ammount):
    ammount = int(ammount)
    if 0 <= ammount <= 1000:
        return "Nominal Member"

    return "Member"


def getMember_Type(customer):
    if customer == "Member":
        return "A Type"

    return "B Type"


def addAllDetails(data):
    print("Adding All Details...")
    i = 1
    j = 1
    vill = data[0]["Village"]
    for x in range(len(data)):
        gender = getGender(data[x]["Surname"])
        Customer_Type = getCustomer_Type(data[x]["Hissa"])
        Member_Type = getMember_Type(Customer_Type)
        farmertype = getFarmer(data[x]["Hissa"])
        data[x]["Gender"] = gender
        data[x]["Farmer"] = farmertype
        data[x]["Member_Type"] = Member_Type
        data[x]["Customer_Type"] = Customer_Type

        if vill == data[x]["Village"]:
            str = f"{i}/{j}"
            data[x]["Ledger"] = str
            j = j + 1
        else:
            vill = data[x]["Village"]
            i = i + 1
            str = f"{i}/1"
            data[x]["Ledger"] = str
            j = 2

    print("Added All Details")
    return data


def addFather(data):
    res = [sub.replace('phaadar', 'father') for sub in data]
    print("Add Father complete")
    return res


def removecharacters(name, data):
    # res = [i.title() for i in data]
    res = data

    res = [sub.replace('aa', 'a') for sub in res]
    res = [sub.replace('ee', 'i') for sub in res]
    res = [sub.replace('oo', 'u') for sub in res]

    # for x in res:
    #     print(x)

    print("Done Characters ", name)
    return res


def readSampleName():
    m_row = get_maximum_rows(namessheet_obj)
    sampleName = []
    for i in range(2, m_row + 1):
        name = namessheet_obj.cell(row=i, column=1).value
        vill = namessheet_obj.cell(row=i, column=2).value
        hissa = namessheet_obj.cell(row=i, column=3).value

        nameobj = {
            "Member": name,
            "Gender": "",
            "Village": vill,
            "Hissa": hissa,
            "Ledger": "",
            "Farmer": "",
            "Surname": "",
            "Father": "",
            "Adhar": ""
        }
        sampleName.append(nameobj)
    print("Done Read Name And Added to Array")
    return sampleName


def readMaleFemaleName():
    print("Reading All Male Female Data....")
    m_row = get_maximum_rows(AllNamesSheet_obj)
    for i in range(1, m_row + 1):
        male = AllNamesSheet_obj.cell(row=i, column=3).value
        female = AllNamesSheet_obj.cell(row=i, column=1).value

        if male is not None:
            AllMale.append(male)
        if female != None:
            AllFemale.append(female)

    print("Done Read All Name MAle Female")


def updateMainData(sheetdata, englishName, englishSurname, englishFather, englishVillage):
    for i in range(len(sheetdata)):
        sheetdata[i]["Member"] = englishName[i].capitalize()
        sheetdata[i]["Surname"] = englishSurname[i].capitalize()
        sheetdata[i]["Father"] = englishFather[i].capitalize()
        sheetdata[i]["Village"] = englishVillage[i]

    return sheetdata


# def formateDataProper(data):
#     for x in data:


#
# print()

# 1 read excel data
# readsheetdata()
# sort data

# sheetdata = sorted(sheetdata, key=itemgetter('Village'))

# 2 excel data to english
# unicodetoenglish()

# 3 remove characters
# englishFather = addFather(englishFather)
# englishSurname = removecharacters(englishSurname)
# englishVillage = removecharacters(englishVillage)
# englishName = removecharacters(englishName)
# englishFather = removecharacters(englishFather)

# 4 update Main Data Add English Name And Village Names
#     updateMainData()

# 5 add All Details Farmer Type Ledger Folio and Gender
#     addAllDetails(sheetdata)

# 6 Save data to excel ouput file

# addheadres()
# savedatatofile()
# readSampleName()
# readMaleFemaleName()

# sampleName = sorted(sampleName, key=itemgetter('Village'))
# print("Sorting Done")
#
# addAllDetails(sampleName)
# savedatatofile(sampleName)

# for x in sampleName:
#     print(x["Member"], x["Gender"], x["Village"], x["Ledger"])

def handle_convert():
    # message_label.config(text="Loading...")
    inpt_obj = openpyxl.load_workbook(inputFile)
    inptsheet_obj = inpt_obj.active
    sheetdata = readsheetdata(inptsheet_obj)
    allEnglishdata = unicodetoenglish(sheetdata)

    englishSurname = removecharacters("Surname", allEnglishdata["Surname"])
    englishVillage = removecharacters("Village", allEnglishdata["Village"])
    englishName = removecharacters("Name", allEnglishdata["Name"])
    englishFather = addFather(allEnglishdata["Father"])
    englishFather = removecharacters("Father", englishFather)

    # for x in englishVillage:
    #     print(x)

    sheetdata = updateMainData(sheetdata, englishName, englishSurname, englishFather, englishVillage)

    sheetdata = sorted(sheetdata, key=itemgetter('Village'))
    sheetdata = addAllDetails(sheetdata)

    savedatatofile(sheetdata)

    message_label.config(text="Converted")
    print("Converted")

     # try:
     #    inpt_obj = openpyxl.load_workbook(inputFile)
     #    inptsheet_obj = inpt_obj.active
     #    sheetdata = readsheetdata(inptsheet_obj)
     #    allEnglishdata = unicodetoenglish(sheetdata)
     #
     #    englishSurname = removecharacters("Surname", allEnglishdata["Surname"])
     #    englishVillage = removecharacters("Village", allEnglishdata["Village"])
     #    englishName = removecharacters("Name", allEnglishdata["Name"])
     #    englishFather = addFather(allEnglishdata["Father"])
     #    englishFather = removecharacters("Father", englishFather)
     #
     #    sheetdata = updateMainData(sheetdata, englishName, englishSurname, englishFather, englishVillage)
     #
     #    sheetdata = sorted(sheetdata, key=itemgetter('Village'))
     #    sheetdata = addAllDetails(sheetdata)
     #
     #    savedatatofile(sheetdata)
     #
     #    message_label.config(text="Converted")
     #    print("Converted")
     #
     # except Exception as error:
     #    print(error)

def handle_reset():
    message_label.config(text="")
    print("Reset")

#
root = Tk()

root.title('DCT FILE')
root.iconbitmap('favicon1.ico')

root.geometry('350x500')

root.configure(background='#0096DC')
img = Image.open('logo.png')
resized_img = img.resize((70, 70))
img = ImageTk.PhotoImage(resized_img)

img_label = Label(root, image=img)
img_label.pack(pady=(10, 10))

text_label = Label(root, text='DCT File Converter', fg='white', bg='#0096DC')
text_label.pack()
text_label.config(font=('verdana', 24))

status_lable = Label(root, text='Ready For Conversion', fg='white', bg='#0096DC')
status_lable.pack(pady=(20, 5))
status_lable.config(font=('verdana', 12))

message_label = Label(root, text='', fg='white', bg='#0096DC')
message_label.pack(pady=(20, 5))
message_label.config(font=('verdana', 12))

login_btn = Button(root, text='Convert Data', bg='#00a65a', fg='white', width=20, height=2, command=handle_convert)
login_btn.pack(pady=(10, 20))
login_btn.config(font=('verdana', 10))

reset_btn = Button(root, text='Reset Data', bg='white', fg='black', width=20, height=2, command=handle_reset)
reset_btn.pack(pady=(10, 20))
reset_btn.config(font=('verdana', 10))

# b1 = ttk.Button(root, text="Reset Data",command=handle_reset)
# b1.pack( padx=5, pady=15)
# b1.config(width=20)

# readSampleName()
readMaleFemaleName()
#
root.mainloop()


#
# inpt_obj = openpyxl.load_workbook(inputFile)
# inptsheet_obj = inpt_obj.active
# sheetdata = readsheetdata(inptsheet_obj)
#
# # savedatatofile(sheetdata)
# line=1
# for x in sheetdata:
#     print(line,x["Member"])
#     line=line+1


# inpt_obj = openpyxl.load_workbook(inputFile)
# inptsheet_obj = inpt_obj.active
# sheetdata = readsheetdata(inptsheet_obj)
# allEnglishdata = unicodetoenglish(sheetdata)

# englishSurname = removecharacters("Surname", allEnglishdata["Surname"])
# englishVillage = removecharacters("Village", allEnglishdata["Village"])
# englishName = removecharacters("Name", allEnglishdata["Name"])
# englishFather = addFather(allEnglishdata["Father"])
# englishFather = removecharacters("Father", englishFather)

# vildata=[]
# for i in sheetdata:
#     vildata.append(i["Village"])
# print("Vill data",len(vildata))

# memdata=[]
# for i in sheetdata:
#     memdata.append(i["Member"])
# print("member",len(memdata))

# for x in allEnglishdata["Village"]:
#     print(x)
# # for x in allEnglishdata["Father"]:
#     print(x)

# newdata=toEnglish(200,vildata)
# striped= list(map(str.strip, newdata.split(',')))
# print(len(striped))
# print(striped)

# for i in range(len(striped)):
#     print(i+1,striped[i])

# memdata=toEnglish(180,memdata)
# striped1= list(map(str.strip, memdata.split(',')))
# print(len(striped1))
# for i in striped:
#     print(i)